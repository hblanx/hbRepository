# -*- coding: utf-8 -*-
import openpyxl
from numpy import random, dot, exp, array, savez, load
from setting import inputList, saveUrl, testList

'''
getHang函数 输入 名字列表 和 excel表 ，返回 姓名:行号 的字典
'''


def getHang(nameList, ws):
    allList = dict([(name, "") for name in nameList])  # 生成名字对应的空字典
    maxHan = ws.max_row  # 获取最大行数
    for i in range(2, maxHan + 1):
        name = ws.cell(row=i, column=2).value
        if name in allList.keys():
            allList[name] = i

    return allList


'''
getDetailScore函数 输入 姓名:行号 的字典 和 excel表 和 目标人名
返回 训练因素numpy矩阵 和 训练目标numpy矩阵
'''


def getDetailScore(allList, ws, aim):
    aimList = []  # 目标矩阵
    otherList = []  # 其他矩阵
    aimRow = allList.pop(aim)
    for i in range(6, 65):  # 处理其他矩阵
        oneLessonList = []
        for j in allList.values():
            oneLessonScore = ws.cell(row=j, column=i).value
            # 数据清洗
            if isReasonable(oneLessonScore):
                oneLessonList.append(float(oneLessonScore) / 100)
            else:
                break
        # 处理目标矩阵并添加数据到最终列表中
        if len(oneLessonList) == 3:
            oneLessonScore = ws.cell(row=aimRow, column=i).value
            if isReasonable(oneLessonScore):
                aimList.append(float(oneLessonScore) / 100)
                otherList.append(oneLessonList)

    otherList = array(otherList)
    aimList = array([aimList]).T
    return otherList, aimList


'''
isReasonable函数 输入excel单元格数据 判断是否合理
合理时 返回 Ture 不合理时 返回 False
'''


def isReasonable(score):
    if score is not None and score != "" and score.isdigit():
        return True
    else:
        return False


'''
learn函数 输入 训练因素numpy矩阵 和 训练目标numpy矩阵 和 保存路径
如果保存路径非False进行保存 返回 w0,w1,w2
'''


def learn(x, y, saveUrl):
    random.seed(1)
    # 神经元数量
    w0 = random.random((3, 35)) * 2 - 1
    w1 = random.random((35, 18)) * 2 - 1
    w2 = random.random((18, 1)) * 2 - 1

    for it in range(50000):
        l0 = x
        l1, l2, l3 = fp(l0, w0, w1, w2)
        l0_delta, l1_delta, l2_delta = bp(l1, l2, l3, w0, w1, w2, y)
        w2 = w2 + dot(l2.T, l2_delta)
        w1 = w1 + dot(l1.T, l1_delta)
        w0 = w0 + dot(l0.T, l0_delta)
    if saveUrl:
        savez(saveUrl, w0=w0, w1=w1, w2=w2)
    return w0, w1, w2


'''
loadWeights函数 输入 保存地址
如果地址为False报错，反之 返回 w0,w1,w2
'''


def loadWeights(saveUrl):
    assert saveUrl, "saveUrl为空，请检查设置"
    weights = load(saveUrl)
    w0 = weights["w0"]
    w1 = weights["w1"]
    w2 = weights["w2"]
    weights.close()
    return w0, w1, w2


'''
train函数 输入 w0,w1,w2元组 和 全局变量testList
打印 概率 无返回值
'''


def train(weights):
    ans = fp(testList, *weights)[2]
    print(f"预测成绩为{int(ans[0] * 100)}")


def fp(l0, w0, w1, w2):
    l1 = 1 / (1 + exp(-dot(l0, w0)))
    l2 = 1 / (1 + exp(-dot(l1, w1)))
    l3 = 1 / (1 + exp(-dot(l2, w2)))
    return l1, l2, l3


def bp(l1, l2, l3, w0, w1, w2, y):
    l2_error = y - l3
    l2_slope = l3 * (1 - l3)
    l2_delta = l2_slope * l2_error

    l1_error = l2_delta.dot(w2.T)
    l1_slope = l2 * (1 - l2)
    l1_delta = l1_slope * l1_error

    l0_error = l1_delta.dot(w1.T)
    l0_slope = l1 * (1 - l1)
    l0_delta = l0_slope * l0_error

    return l0_delta, l1_delta, l2_delta


if __name__ == "__main__":
    # 注意此处修改为自己的数据
    wb = openpyxl.load_workbook('myexcel.xlsx')
    ws1 = wb["sheet1"]
    allList = getHang(inputList, ws1)
    otherList, aimList = getDetailScore(allList, ws1, inputList[0])
    testWeights = learn(otherList, aimList, saveUrl)
    # rr=loadWeights(saveUrl)
    train(testWeights)
