# -*- coding: utf-8 -*-
import openpyxl
from numpy import random, dot, exp, array, savez, load


class hbscore:
    def __init__(self, saveUrl="", nameList=[], testList=[]):
        if len(testList) == 3:
            for i in range(0,3):
                testList[i] = round(testList[i] / 100,2)
        self.__saveUrl = saveUrl
        self.__nameList = nameList
        self.__testList = testList
        self.__otherList = ""
        self.__aimList = ""
        self.__weights = ""

    def setSaveUrl(self, saveUrl):
        self.__saveUrl = saveUrl

    def setNameList(self, nameList):
        self.__nameList = nameList

    def setTestList(self, testList):
        self.__testList = testList

    '''
    excelToArr函数 接收setting字典，键包含excelName和sheetName
    '''

    def excelToArr(self, setting):
        # 打开excel
        try:
            wb = openpyxl.load_workbook(setting["excelName"])
            ws = wb[setting["sheetName"]]
        except:
            raise ValueError("excel设置错误")
        # 生成 名字：行号 字典
        allList = dict([(name, "") for name in self.__nameList])  # 生成名字对应的空字典
        maxHan = ws.max_row  # 获取最大行数
        for i in range(2, maxHan + 1):
            name = ws.cell(row=i, column=2).value
            if name in allList.keys():
                allList[name] = i
        wb.close()
        # 获取详细成绩并生成numpy数组
        aimList = []  # 目标矩阵
        otherList = []  # 其他矩阵
        aimRow = allList.pop(self.__nameList[0])
        for i in range(6, 65):  # 处理其他矩阵
            oneLessonList = []
            for j in allList.values():
                oneLessonScore = ws.cell(row=j, column=i).value
                # 数据清洗
                if self._isReasonable(oneLessonScore):
                    oneLessonList.append(float(oneLessonScore) / 100)
                else:
                    break
            # 处理目标矩阵并添加数据到最终列表中
            if len(oneLessonList) == 3:
                oneLessonScore = ws.cell(row=aimRow, column=i).value
                if self._isReasonable(oneLessonScore):
                    aimList.append(float(oneLessonScore) / 100)
                    otherList.append(oneLessonList)

        self.__otherList = array(otherList)
        self.__aimList = array([aimList]).T

    def _isReasonable(self, score):
        if score is not None and score != "" and score.isdigit():
            return True
        else:
            return False

    def learn(self):
        random.seed(1)
        # 神经元数量
        w0 = random.random((3, 35)) * 2 - 1
        w1 = random.random((35, 18)) * 2 - 1
        w2 = random.random((18, 1)) * 2 - 1

        for it in range(50000):
            l0 = self.__otherList
            l1, l2, l3 = self.__fp(l0, w0, w1, w2)
            l0_delta, l1_delta, l2_delta = self.__bp(l1, l2, l3, w1, w2, self.__aimList)
            w2 = w2 + dot(l2.T, l2_delta)
            w1 = w1 + dot(l1.T, l1_delta)
            w0 = w0 + dot(l0.T, l0_delta)
        self.__weights = (w0, w1, w2)

    def __fp(self, l0, w0, w1, w2):
        l1 = 1 / (1 + exp(-dot(l0, w0)))
        l2 = 1 / (1 + exp(-dot(l1, w1)))
        l3 = 1 / (1 + exp(-dot(l2, w2)))
        return l1, l2, l3

    def __bp(self, l1, l2, l3, w1, w2, y):
        l2_error = y - l3
        l2_slope = l3 * (1 - l3)
        l2_delta = l2_slope * l2_error

        l1_error = l2_delta.dot(w2.T)
        l1_slope = l2 * (1 - l2)
        l1_delta = l1_slope * l1_error

        l0_error = l1_delta.dot(w1.T)
        l0_slope = l1 * (1 - l1)
        l0_delta = l0_slope * l0_error

        return l0_delta, l1_delta, l2_delta

    def save(self):
        try:
            savez(self.__saveUrl, w0=self.__weights[0], w1=self.__weights[1], w2=self.__weights[2])
        except:
            raise Exception("保存权重文件时出错")

    def detect(self):
        if len(self.__testList) == 3:
            for i in range(0, 3):
                self.__testList[i] = round(self.__testList[i] / 100, 2)
        ans = self.__fp(self.__testList, *self.__weights)[2]
        return int(ans[0] * 100) // 1

    def load(self):
        try:
            weights = load(self.__saveUrl)
            w0 = weights["w0"]
            w1 = weights["w1"]
            w2 = weights["w2"]
            weights.close()
            self.__weights = (w0, w1, w2)
        except:
            raise Exception("读取权重文件时出错")

if __name__ == "__main__":
    myhbscore = hbscore(saveUrl="./weights.npz", nameList=["zh", "gg", "zh", "sz"], testList=[])
    mysetting = {"excelName": "myexcel.xlsx", "sheetName": "sheet1"}
    myhbscore.excelToArr(mysetting)
    myhbscore.learn()
    myhbscore.setTestList([79, 67, 80])
    #myhbscore.setSaveUrl()
    #myhbscore.setNameList()
    myhbscore.save()
    myhbscore.load()
    print(myhbscore.detect())
